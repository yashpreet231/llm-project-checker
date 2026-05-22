"""
Classroom routes — classes, materials, quizzes, polls, announcements,
and teacher evaluation of student submissions.
"""
from __future__ import annotations

from datetime import datetime
from typing import List, Optional
import uuid

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app import classroom as cls
from app.auth import UserRecord, current_user, require_role
from app.persistence import save_store
from app.utils.file_parser import extract_text

router = APIRouter()


# ── request/response models ─────────────────────────────────────��─────────────

class ClassroomOut(BaseModel):
    id:           str
    subject:      str
    description:  str
    teacher_name: str
    class_code:   str
    student_count: int
    enrolled:     bool


class CreateClassBody(BaseModel):
    subject:     str = Field(min_length=1, max_length=120)
    description: str = Field(max_length=500, default="")
    class_code:  str = Field(min_length=2, max_length=20)


class JoinClassBody(BaseModel):
    class_code: str = Field(min_length=1)


class MaterialBody(BaseModel):
    title: str = Field(min_length=1, max_length=200)
    text:  str = Field(min_length=1)


class GenerateQuizBody(BaseModel):
    difficulty: str = "medium"
    count:      int = 5
    title:      Optional[str] = None
    due_date:   Optional[str] = None


class SubmitQuizBody(BaseModel):
    answers: List[str]


class EvaluateBody(BaseModel):
    remarks: str = Field(min_length=1)
    score:   int = Field(ge=0, le=100)


class MessageBody(BaseModel):
    title: str = Field(min_length=1, max_length=200)
    body:  str = Field(min_length=1, max_length=2000)


class CreatePollBody(BaseModel):
    question: str = Field(min_length=1, max_length=300)
    options:  List[str] = Field(min_length=2, max_length=6)


class VotePollBody(BaseModel):
    option_index: int = Field(ge=0)


class ChatBody(BaseModel):
    message: str = Field(min_length=1, max_length=2000)


class VoiceChatBody(BaseModel):
    audio_b64:     str = Field(min_length=1)
    language_code: str = Field(default="hi-IN")


class CreateProjectBody(BaseModel):
    title:    str = Field(min_length=1, max_length=200)
    description: str = Field(min_length=1)
    due_date: Optional[str] = None


class SubmitProjectBody(BaseModel):
    github_url: str = Field(min_length=1)


class EvalProjectBody(BaseModel):
    remarks: str = Field(min_length=1)
    score:   int = Field(ge=0, le=100)


# ── helpers ────────────────��──────────────────────────────────────────────────

def _as_out(c: cls.Classroom, user: UserRecord) -> ClassroomOut:
    return ClassroomOut(
        id=c.id,
        subject=c.subject,
        description=c.description,
        teacher_name=c.teacher_name,
        class_code=c.class_code,
        student_count=len(c.student_ids),
        enrolled=(user.id in c.student_ids) or (user.id == c.teacher_id),
    )


def _require_classroom(cid: str) -> cls.Classroom:
    c = cls._classrooms.get(cid)
    if not c:
        raise HTTPException(status_code=404, detail="Classroom not found")
    return c


def _require_owner(c: cls.Classroom, user: UserRecord):
    if c.teacher_id != user.id:
        raise HTTPException(status_code=403, detail="Not your classroom.")


# ── fixed-path routes (before /{cid} wildcard) ───────────────────────────────

@router.get("", response_model=List[ClassroomOut])
def list_classes(user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    out = []
    for c in cls._classrooms.values():
        if user.role == "teacher" and c.teacher_id != user.id:
            continue
        if user.role == "student" and user.id not in c.student_ids:
            continue
        out.append(_as_out(c, user))
    return out


@router.post("", dependencies=[Depends(require_role("teacher"))])
def create_class(body: CreateClassBody, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    for c in cls._classrooms.values():
        if c.class_code.lower() == body.class_code.strip().lower():
            raise HTTPException(status_code=409, detail="A class with that code already exists.")
    c = cls.Classroom(
        id=str(uuid.uuid4()),
        subject=body.subject.strip(),
        description=body.description.strip(),
        teacher_id=user.id,
        teacher_name=user.name,
        class_code=body.class_code.strip().upper(),
    )
    cls._classrooms[c.id] = c
    save_store()
    return _as_out(c, user)


@router.post("/join")
def join_class(body: JoinClassBody, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    if user.role != "student":
        raise HTTPException(status_code=403, detail="Only students can join classes.")
    try:
        c = cls.join_class(body.class_code, user.id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return _as_out(c, user)


@router.post("/assignments/{quiz_id}/submit")
def submit_quiz_answers(quiz_id: str, body: SubmitQuizBody,
                        user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        sub = cls.submit_quiz(quiz_id, user.id, user.name, body.answers)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Evaluation error: {e}")
    return sub


@router.post("/submissions/{sid}/evaluate",
             dependencies=[Depends(require_role("teacher"))])
def evaluate_submission(sid: str, body: EvaluateBody,
                        user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    s = cls._submissions.get(sid)
    if not s:
        raise HTTPException(status_code=404, detail="Submission not found")
    c = cls._classrooms.get(s.classroom_id)
    if not c or c.teacher_id != user.id:
        raise HTTPException(status_code=403, detail="Not your classroom.")
    updated = s.model_copy(update={
        "remarks":      body.remarks,
        "score":        body.score,
        "evaluated_at": datetime.utcnow(),
        "evaluated_by": user.name,
    })
    cls._submissions[sid] = updated
    save_store()
    return updated


@router.post("/polls/{pid}/vote")
def vote_poll(pid: str, body: VotePollBody,
              user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    poll = cls._polls.get(pid)
    if not poll:
        raise HTTPException(status_code=404, detail="Poll not found")
    if user.id in poll.voters:
        raise HTTPException(status_code=409, detail="You already voted.")
    if body.option_index >= len(poll.options):
        raise HTTPException(status_code=400, detail="Invalid option.")
    poll.options[body.option_index].votes += 1
    poll.voters.append(user.id)
    cls._polls[pid] = poll
    save_store()
    return poll


# ── project fixed-path routes ────────────────────────────────────────────────

@router.post("/projects/{pid}/submit")
def submit_project_route(pid: str, body: SubmitProjectBody,
                         user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        ps = cls.submit_project(pid, user.id, user.name, body.github_url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ps


@router.get("/projects/{pid}/submissions",
            dependencies=[Depends(require_role("teacher"))])
def list_project_submissions(pid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    proj = cls._projects.get(pid)
    if not proj:
        raise HTTPException(status_code=404, detail="Project not found")
    c = cls._classrooms.get(proj.classroom_id)
    if not c or c.teacher_id != user.id:
        raise HTTPException(status_code=403, detail="Not your project.")
    subs = [ps for ps in cls._project_submissions.values() if ps.project_id == pid]
    return sorted(subs, key=lambda s: s.submitted_at, reverse=True)


@router.post("/project-submissions/{psid}/check-progress")
def check_progress_route(psid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        ps = cls.check_github_progress(psid)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ps


@router.post("/project-submissions/{psid}/evaluate-ai")
def evaluate_project_ai_route(psid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        ps = cls.evaluate_project_ai(psid)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ps


@router.post("/project-submissions/{psid}/evaluate",
             dependencies=[Depends(require_role("teacher"))])
def evaluate_project_teacher(psid: str, body: EvalProjectBody,
                             user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    ps = cls._project_submissions.get(psid)
    if not ps:
        raise HTTPException(status_code=404, detail="Submission not found")
    updated = ps.model_copy(update={
        "score": body.score,
        "remarks": body.remarks,
        "evaluated_at": datetime.utcnow(),
        "evaluated_by": user.name,
    })
    cls._project_submissions[psid] = updated
    save_store()
    return updated


@router.post("/project-submissions/{psid}/weekly-snapshot")
def weekly_snapshot_route(psid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        ps = cls.record_weekly_snapshot(psid)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ps


class ProjectChatBody(BaseModel):
    message: str = Field(min_length=1, max_length=2000)


@router.get("/projects/{pid}/chat")
def get_project_chat(pid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    history = cls.get_project_chat_history(pid, user.id)
    return [m.model_dump() for m in history]


@router.post("/projects/{pid}/chat")
def post_project_chat(pid: str, body: ProjectChatBody,
                      user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    try:
        reply = cls.project_guidance_chat(pid, user.id, body.message.strip())
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Guidance error: {e}")
    return reply.model_dump()


# ── parameterised /{cid} routes ─────────────────────────────────────────────

@router.get("/{cid}")
def get_class(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    return {
        **_as_out(c, user).model_dump(),
        "materials": [m.model_dump() for m in c.materials],
        "is_owner":  c.teacher_id == user.id,
    }


# ── stream (announcements feed) ──────────────────────────────────────���───────

@router.get("/{cid}/stream")
def get_stream(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    items = sorted(
        [a for a in cls._announcements.values() if a.classroom_id == cid],
        key=lambda a: a.created_at,
        reverse=True,
    )
    result = []
    for a in items:
        d = a.model_dump()
        if a.kind == "poll" and a.ref_id:
            poll = cls._polls.get(a.ref_id)
            if poll:
                d["poll"] = poll.model_dump()
                d["poll"]["voted"] = user.id in poll.voters
        if a.kind == "quiz" and a.ref_id:
            quiz = cls._quizzes.get(a.ref_id)
            if quiz:
                d["quiz_info"] = {"id": quiz.id, "title": quiz.title, "difficulty": quiz.difficulty, "question_count": len(quiz.questions)}
        result.append(d)
    return result


# ── messages (teacher broadcasts) ────────────────────────────────────────────

@router.post("/{cid}/messages", dependencies=[Depends(require_role("teacher"))])
def post_message(cid: str, body: MessageBody,
                 user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    ann = cls._post_announcement(cid, "message", body.title.strip(), body.body.strip(), author_name=user.name)
    return ann


# ── materials ────────��─────────────────────────────���─────────────────────────

@router.post("/{cid}/materials", dependencies=[Depends(require_role("teacher"))])
def upload_material(cid: str, body: MaterialBody,
                    user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    item = cls.MaterialItem(
        id=str(uuid.uuid4()),
        title=body.title.strip(),
        text=body.text.strip(),
    )
    updated = c.model_copy(update={"materials": [*c.materials, item]})
    cls._classrooms[cid] = updated
    cls._post_announcement(cid, "material", f"New material: {item.title}",
                           body=item.text[:200] + ("..." if len(item.text) > 200 else ""),
                           ref_id=item.id, author_name=user.name)
    return item


@router.post("/{cid}/materials/upload", dependencies=[Depends(require_role("teacher"))])
async def upload_material_file(
    cid: str,
    title: str = Form(...),
    file: UploadFile = File(...),
    user: UserRecord = Depends(current_user),
):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10 MB).")

    text = extract_text(file.filename or "file.txt", content)
    if not text.strip():
        raise HTTPException(status_code=422, detail="Could not extract any text from this file.")

    mat_title = title.strip() or file.filename or "Untitled"
    item = cls.MaterialItem(
        id=str(uuid.uuid4()),
        title=mat_title,
        text=text.strip(),
    )
    updated = c.model_copy(update={"materials": [*c.materials, item]})
    cls._classrooms[cid] = updated
    cls._post_announcement(cid, "material", f"New material: {item.title}",
                           body=item.text[:200] + ("..." if len(item.text) > 200 else ""),
                           ref_id=item.id, author_name=user.name)
    return item


@router.delete("/{cid}/materials/{mid}",
               dependencies=[Depends(require_role("teacher"))])
def delete_material(cid: str, mid: str,
                    user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    remaining = [m for m in c.materials if m.id != mid]
    if len(remaining) == len(c.materials):
        raise HTTPException(status_code=404, detail="Material not found")
    cls._classrooms[cid] = c.model_copy(update={"materials": remaining})
    save_store()
    return {"deleted": mid}


# ── quizzes ──────────────────────────────────────────────────────────────────

@router.post("/{cid}/quiz", dependencies=[Depends(require_role("teacher"))])
def teacher_generate_quiz(cid: str, body: GenerateQuizBody,
                          user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    title = body.title or f"Assignment — {c.subject} ({body.difficulty})"
    due_dt = None
    if body.due_date:
        try:
            due_dt = datetime.fromisoformat(body.due_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid due_date format. Use ISO 8601.")
    try:
        quiz = cls.generate_quiz(cid, body.difficulty, body.count, kind="assignment", title=title, due_date=due_dt)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"LLM error: {e}")
    cls._post_announcement(cid, "quiz", f"New quiz: {quiz.title}",
                           body=f"{len(quiz.questions)} questions · {quiz.difficulty} difficulty",
                           ref_id=quiz.id, author_name=user.name)
    return quiz


@router.post("/{cid}/practice")
def student_practice_quiz(cid: str, body: GenerateQuizBody,
                          user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    title = body.title or "Practice quiz"
    try:
        quiz = cls.generate_quiz(cid, body.difficulty, body.count, kind="practice", title=title)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"LLM error: {e}")
    return quiz


# ── polls ────────────────────────────────────────────────────────────────────

@router.post("/{cid}/polls", dependencies=[Depends(require_role("teacher"))])
def create_poll(cid: str, body: CreatePollBody,
                user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    poll = cls.Poll(
        id=str(uuid.uuid4()),
        classroom_id=cid,
        question=body.question.strip(),
        options=[cls.PollOption(label=o.strip()) for o in body.options],
    )
    cls._polls[poll.id] = poll
    cls._post_announcement(cid, "poll", f"Poll: {poll.question}",
                           ref_id=poll.id, author_name=user.name)
    return poll


# ── assignments & submissions ──────────────────────────────────────────��─────

@router.get("/{cid}/assignments")
def list_assignments(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    return sorted(
        [q for q in cls._quizzes.values()
         if q.classroom_id == cid and q.kind == "assignment"],
        key=lambda q: q.created_at,
        reverse=True,
    )


@router.get("/{cid}/my-submissions")
def my_submissions(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    return sorted(
        [s for s in cls._submissions.values()
         if s.classroom_id == cid and s.student_id == user.id],
        key=lambda s: s.submitted_at,
        reverse=True,
    )


@router.get("/{cid}/submissions", dependencies=[Depends(require_role("teacher"))])
def list_submissions(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    return sorted(
        [s for s in cls._submissions.values() if s.classroom_id == cid],
        key=lambda s: s.submitted_at,
        reverse=True,
    )


# ── projects (GitHub-tracked) ───────────────────────────────────────────────

@router.get("/{cid}/projects")
def list_projects(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    projs = sorted(
        [p for p in cls._projects.values() if p.classroom_id == cid],
        key=lambda p: p.created_at, reverse=True,
    )
    result = []
    for p in projs:
        d = p.model_dump()
        if user.role == "student":
            my_sub = next((ps for ps in cls._project_submissions.values()
                           if ps.project_id == p.id and ps.student_id == user.id), None)
            d["my_submission"] = my_sub.model_dump() if my_sub else None
        else:
            sub_count = sum(1 for ps in cls._project_submissions.values() if ps.project_id == p.id)
            d["submission_count"] = sub_count
        result.append(d)
    return result


@router.post("/{cid}/projects", dependencies=[Depends(require_role("teacher"))])
def create_project_json(cid: str, body: CreateProjectBody,
                        user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)
    due_dt = None
    if body.due_date:
        try:
            due_dt = datetime.fromisoformat(body.due_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid due_date.")
    p = cls.create_project(cid, body.title.strip(), body.description.strip(), due_date=due_dt)
    return p


@router.post("/{cid}/projects/upload", dependencies=[Depends(require_role("teacher"))])
async def create_project_file(
    cid: str,
    title: str = Form(...),
    due_date: str = Form(default=""),
    file: UploadFile = File(...),
    user: UserRecord = Depends(current_user),
):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (max 10 MB).")

    text = extract_text(file.filename or "project.txt", content)
    if not text.strip():
        raise HTTPException(status_code=422, detail="Could not extract text from this file.")

    due_dt = None
    if due_date:
        try:
            due_dt = datetime.fromisoformat(due_date)
        except ValueError:
            pass

    p = cls.create_project(
        cid, title.strip(), text.strip(),
        file_name=file.filename, due_date=due_dt,
    )
    return p


@router.get("/{cid}/my-project-submissions")
def my_project_submissions(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    return sorted(
        [ps for ps in cls._project_submissions.values()
         if ps.classroom_id == cid and ps.student_id == user.id],
        key=lambda ps: ps.submitted_at, reverse=True,
    )


@router.get("/{cid}/leaderboard")
def get_leaderboard(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    return cls.get_leaderboard(cid)


# ── AI tutor chat ───────────────────────────────────────────────────────────

@router.get("/{cid}/chat")
def get_chat(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    history = cls.get_chat_history(cid, user.id)
    return [m.model_dump() for m in history]


@router.get("/{cid}/analytics")
def get_analytics(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    is_teacher = c.teacher_id == user.id

    if is_teacher:
        subs = [s for s in cls._submissions.values() if s.classroom_id == cid and s.score is not None]
    else:
        subs = [s for s in cls._submissions.values() if s.classroom_id == cid and s.student_id == user.id and s.score is not None]

    if not subs:
        return {
            "scores": [],
            "average": None,
            "best": None,
            "worst": None,
            "total_submissions": 0,
            "class_average": None,
            "student_count": len(c.student_ids),
        }

    subs_sorted = sorted(subs, key=lambda s: s.submitted_at)
    scores = [{"title": s.title, "score": s.score, "date": s.submitted_at.isoformat(),
               "student_name": s.student_name, "correct": s.correct_count, "total": s.total_count} for s in subs_sorted]
    avg = round(sum(s.score for s in subs) / len(subs), 1)
    best = max(s.score for s in subs)
    worst = min(s.score for s in subs)

    all_subs = [s for s in cls._submissions.values() if s.classroom_id == cid and s.score is not None]
    class_avg = round(sum(s.score for s in all_subs) / len(all_subs), 1) if all_subs else None

    return {
        "scores": scores,
        "average": avg,
        "best": best,
        "worst": worst,
        "total_submissions": len(subs),
        "class_average": class_avg,
        "student_count": len(c.student_ids),
    }


@router.post("/{cid}/chat")
def post_chat(cid: str, body: ChatBody, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    try:
        reply = cls.chat_with_tutor(cid, user.id, body.message.strip())
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Tutor error: {e}")
    return reply.model_dump()


# ── voice chat (Sarvam AI) ──────────────────────────────────────────────────

@router.get("/{cid}/voice/languages")
def voice_languages(cid: str, user: UserRecord = Depends(current_user)):
    from app.services.sarvam import SUPPORTED_LANGUAGES, is_configured
    _require_classroom(cid)
    return {"configured": is_configured(), "languages": SUPPORTED_LANGUAGES}


@router.post("/{cid}/voice")
async def voice_chat(cid: str, body: VoiceChatBody,
                     user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    _require_classroom(cid)
    from app.services.sarvam import voice_tutor_pipeline, is_configured
    if not is_configured():
        raise HTTPException(status_code=503, detail="Sarvam AI is not configured. Set SARVAM_API_KEY in .env")
    try:
        result = await voice_tutor_pipeline(body.audio_b64, cid, user.id, body.language_code)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Voice pipeline error: {e}")
    if "error" in result:
        raise HTTPException(status_code=422, detail=result["error"])
    return result


# ── lectures + attendance summaries ─────────────────────────────────────────────

class SubmitSummaryBody(BaseModel):
    summary: str = Field(min_length=1, max_length=10000)


class OverrideAttendanceBody(BaseModel):
    granted: bool


@router.get("/{cid}/lectures")
def list_lectures(cid: str, user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    lectures = cls.list_lectures(cid)
    is_teacher = (c.teacher_id == user.id)
    out = []
    for lec in lectures:
        my_sub = cls.get_my_summary(lec.id, user.id)
        out.append({
            "id":             lec.id,
            "title":          lec.title,
            "file_name":      lec.file_name,
            "threshold":      lec.threshold,
            "held_at":        lec.held_at,
            "created_at":     lec.created_at,
            # transcript only visible to the teacher
            "transcript":     lec.transcript if is_teacher else None,
            "my_submission":  my_sub.model_dump() if my_sub else None,
            "submission_count": sum(
                1 for s in cls._summary_submissions.values() if s.lecture_id == lec.id
            ),
        })
    return out


@router.post("/{cid}/lectures",
             dependencies=[Depends(require_role("teacher"))])
async def create_lecture(
    cid: str,
    title:     str = Form(...),
    threshold: int = Form(default=6),
    held_at:   str = Form(default=""),
    transcript_text: str = Form(default=""),
    file: Optional[UploadFile] = File(default=None),
    user: UserRecord = Depends(current_user),
):
    """Teacher creates a lecture. Provide either an uploaded transcript file OR transcript_text."""
    cls._seed_if_empty()
    c = _require_classroom(cid)
    _require_owner(c, user)

    transcript = transcript_text.strip()
    fname = None
    if file is not None:
        content = await file.read()
        if len(content) > 10 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large (max 10 MB).")
        transcript = extract_text(file.filename or "transcript.txt", content).strip()
        fname = file.filename

    if not transcript:
        raise HTTPException(status_code=422, detail="Transcript is empty.")

    held_dt = None
    if held_at:
        try:
            held_dt = datetime.fromisoformat(held_at)
        except ValueError:
            pass

    lec = cls.create_lecture(
        cid, title.strip(), transcript,
        threshold=threshold, file_name=fname, held_at=held_dt,
    )
    return lec


@router.delete("/{cid}/lectures/{lid}",
               dependencies=[Depends(require_role("teacher"))])
def delete_lecture(cid: str, lid: str, user: UserRecord = Depends(current_user)):
    c = _require_classroom(cid)
    _require_owner(c, user)
    lec = cls._lectures.get(lid)
    if not lec or lec.classroom_id != cid:
        raise HTTPException(status_code=404, detail="Lecture not found")
    cls._lectures.pop(lid, None)
    # cascade summaries
    to_remove = [sid for sid, s in cls._summary_submissions.items() if s.lecture_id == lid]
    for sid in to_remove:
        cls._summary_submissions.pop(sid, None)
    save_store()
    return {"deleted": lid, "summaries_removed": len(to_remove)}


@router.post("/{cid}/lectures/{lid}/summary")
def submit_summary(cid: str, lid: str, body: SubmitSummaryBody,
                   user: UserRecord = Depends(current_user)):
    cls._seed_if_empty()
    c = _require_classroom(cid)
    if user.id == c.teacher_id:
        raise HTTPException(status_code=403, detail="Teachers don't submit summaries.")
    if user.id not in c.student_ids:
        raise HTTPException(status_code=403, detail="You are not enrolled in this class.")
    lec = cls._lectures.get(lid)
    if not lec or lec.classroom_id != cid:
        raise HTTPException(status_code=404, detail="Lecture not found")

    try:
        sub = cls.submit_summary(lid, user.id, user.name, body.summary.strip())
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    return sub


@router.get("/{cid}/lectures/{lid}/summaries",
            dependencies=[Depends(require_role("teacher"))])
def list_summaries(cid: str, lid: str, user: UserRecord = Depends(current_user)):
    c = _require_classroom(cid)
    _require_owner(c, user)
    lec = cls._lectures.get(lid)
    if not lec or lec.classroom_id != cid:
        raise HTTPException(status_code=404, detail="Lecture not found")
    return cls.list_summaries(lid)


@router.post("/{cid}/summary-submissions/{ssid}/override",
             dependencies=[Depends(require_role("teacher"))])
def override_attendance(cid: str, ssid: str, body: OverrideAttendanceBody,
                        user: UserRecord = Depends(current_user)):
    c = _require_classroom(cid)
    _require_owner(c, user)
    sub = cls._summary_submissions.get(ssid)
    if not sub or sub.classroom_id != cid:
        raise HTTPException(status_code=404, detail="Summary submission not found")
    return cls.override_attendance(ssid, body.granted, user.id)


@router.get("/{cid}/attendance",
            dependencies=[Depends(require_role("teacher"))])
def attendance_report(cid: str, user: UserRecord = Depends(current_user)):
    c = _require_classroom(cid)
    _require_owner(c, user)
    return cls.attendance_report(cid)


@router.get("/{cid}/lectures/{lid}/attendance-excel",
            dependencies=[Depends(require_role("teacher"))])
def download_attendance_excel(cid: str, lid: str,
                              user: UserRecord = Depends(current_user)):
    """Download an Excel sheet for a single lecture: name, roll no, remarks, attendance."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    from app import auth as authlib

    c = _require_classroom(cid)
    _require_owner(c, user)
    lec = cls._lectures.get(lid)
    if not lec or lec.classroom_id != cid:
        raise HTTPException(status_code=404, detail="Lecture not found")

    # index submissions by student_id for quick lookup
    subs_by_student = {
        s.student_id: s
        for s in cls._summary_submissions.values()
        if s.lecture_id == lid
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── header row ────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")  # blue
    headers = ["#", "Student Name", "Roll No", "Submitted", "AI Score (0-10)",
               "Remarks", "Attendance"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 55
    ws.column_dimensions["G"].width = 14

    green_fill = PatternFill(fill_type="solid", fgColor="D1FAE5")
    red_fill   = PatternFill(fill_type="solid", fgColor="FEE2E2")

    # ── one row per enrolled student ─────────────────────────────────────────
    for row_idx, student_id in enumerate(c.student_ids, start=2):
        entry   = authlib._users_by_id.get(student_id)
        name    = entry["record"].name    if entry else student_id
        roll_no = entry["record"].roll_no if entry else ""
        sub     = subs_by_student.get(student_id)

        submitted   = sub.submitted_at.strftime("%Y-%m-%d %H:%M") if sub else "—"
        ai_score    = sub.ai_score    if sub else None
        remarks     = sub.ai_feedback if sub else "No submission"
        granted     = sub.attendance_granted if sub else False
        attend_text = "Present" if granted else "Absent"
        row_fill    = green_fill if granted else red_fill

        values = [row_idx - 1, name, roll_no, submitted,
                  ai_score, remarks, attend_text]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=(col == 6), vertical="top")

    # freeze the header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_title = lec.title.replace(" ", "_").replace("/", "-")[:60]
    filename = f"attendance_{safe_title}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
